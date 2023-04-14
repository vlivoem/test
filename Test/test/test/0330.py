# openpyxl：编辑excel的模块
import openpyxl
# openpxl.styles：调整excel样式的模块
from openpyxl.styles import Font, colors, Alignment
import xlsxwriter



def update():
    # 获取文件
    ff = openpyxl.load_workbook("D:\\1658369729702.xlsx")
    # 获取工作表
    data = ff["Sheet1"]
    # print(type(data))
    # 按行获取数据
    row_data = list(data.rows)
    # 按列读取数据
    # clumns_data = data.clumns
    # 筛选、输出数据
    sh2 = ff.create_sheet("筛选")
    for device in row_data:
        list_data = []
        for key in device[1],device[2],device[6]:
            list_data.append(key.value)
        sh2.append(list_data)

    sh2.sheet_properties.tabColor = "FFAA33"

    # 创建样式
    font = Font(name='等线', size=13, color=colors.BLACK, bold=True)
    # 创建对齐方式
    alig = Alignment(horizontal='center', vertical='center')
    # 使用样式
    for row in row_data:
        for clumn in row:
            if clumn == row_data[0]:
                clumn.font = font
            clumn.alignment = alig
    filename = "D:\\筛选6.xlsx"
    ff.save(filename)

    # 设置行高列宽
    file2 = xlsxwriter.Workbook("11")
    ff= file2.add_worksheet("11")
    ff.set_column('A:G', 20)
    ff.set_row(3, 20)
    file2.close()







if __name__ == "__main__":
    update()
else:
    print("bye")




